import os
from openpyxl import load_workbook

def make_passport(
        surname_passport_number,
        surname, name, person_number, birthday,
        gender, city,
        country_name_english,
        country_name_local,
        nationality_english,
        nationality_local,
        country_letter):
    workbook = load_workbook("passport_template.xlsx")    
    sheet = workbook.active
    
    surname_cell = sheet.cell(row=7, column=4)
    name_cell = sheet.cell(row=7, column=6)
    gender_cell = sheet.cell(row=13, column=2) 

    birthday_cell = sheet.cell(row=10, column=4) 
    person_number_cell = sheet.cell(row=13, column=6) 

    city_cell = sheet.cell(row=10, column=6)
    country_name_local_cell = sheet.cell(row=4, column=4)
    country_name_english_cell = sheet.cell(row=4, column=5)
    country_letter_cell = sheet.cell(row=8, column=2)
    nationality_cell = sheet.cell(row=13, column=4)


    surname_cell.value = surname
    name_cell.value = name
    gender_cell.value = gender

    person_number_cell.value = person_number
    birthday_cell.value = birthday 
    
    city_cell.value = city
    nationality_cell.value = nationality_local + " / " + nationality_english
    country_name_english_cell.value = country_name_english
    country_name_local_cell.value = country_name_local
    country_letter_cell.value = country_letter

    workbook.save("pass/{}{}.xlsx".format(surname, surname_passport_number))


if __name__ == "__main__":
    make_passport(
        surname_passport_number=0,
        surname="Mia", name="Test", person_number="123",
        birthday="12.12.12",
        gender="male", city="Lillehammer", country_name_english="norway",
        country_name_local="norge", nationality_english="norwegian",
        nationality_local="norsk", country_letter="N")
	

	

		 
